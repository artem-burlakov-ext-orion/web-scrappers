from proxies import *
from bs4 import BeautifulSoup as bs
import requests
import time
from openpyxl import Workbook
import random

pattern = r'job_\d+'
d = []

def get_html(url):
    while True:
        header = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebkit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36'}
        proxy = get_proxy()
        print(proxy)
        try:
            time.sleep(random.choice([5,6,7,8,9,10,11,12,13,14]))
            r = requests.get(url, headers = header, proxies = proxy, timeout = 20)
            return r.text
            break
        except:
            continue

def get_data(html):
    soup = bs(html,'lxml')
    tag = soup.find('div', class_='tag')
    if not tag:
        return  None
    skill = tag.text
    return skill

    # jobs = soup.find('div', id='jobs_list').find_all('div', id=re.compile(pattern))
    # for job in jobs:
    #     time.sleep(4)
    #     specs = job.find_all('a', class_='skill')
    #     for spec in specs[1:]:
    #         skill = spec.text
    #         if skill in ('Старший (Senior)', 'Младший (Junior)', 'Средний (Middle)', 'Стажёр (Intern)'):
    #             continue
    #         d.append(skill)


def get_xlsx(d):
    print('xlsx')
    print(d)
    wb = Workbook()
    ws = wb.active
    # most_common =  Counter(d).most_common(6)
    for i in d:
        m_row = ws.max_row
        ws[f'A{m_row+1}'] = i
        ws[f'B{m_row+1}'] = d[i]
        # ws[f'A{i}'] = str(most_common[i])
    wb.save('moikrug_skills.xlsx')

def main():
    skill_dict = {}
    for i in range(1, 1200):
        time.sleep(10)
        url = f'https://moikrug.ru/vacancies?skills%5B%5D={i}&type=all'
        skill = get_data(get_html(url))
        if skill:
            skill_dict[skill] = i
        print(skill)
    # with open('skills.json', 'w') as file:
    #     json.dump(skill_dict, file)
    get_xlsx(skill_dict)


    # url = 'https://moikrug.ru/vacancies?skills%5B%5D=446&type=all'
    # get_data(get_html(url))
    # for i in range(2,6):
    #     try:
    #         url = f'https://moikrug.ru/vacancies?page={i}&skills%5B%5D=446&type=all'
    #         get_data(get_html(url))
    #     except:
    #         break
    # get_xlsx(d)

if __name__ == '__main__':
    main()