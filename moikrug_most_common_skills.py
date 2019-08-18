from proxies import *
from bs4 import BeautifulSoup as bs
import requests
import time
from openpyxl import Workbook
import random
import re
from collections import Counter

pattern = r'job_\d+'
skills_list = []
major_skills_dict = {'javascript': 264, 'java': 1012, 'python': 446, 'golang': 101, 'postgresql': 537}

myskill = 'python'
myskill_id = major_skills_dict[myskill]

def get_html(url):
    while True:
        header = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebkit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36'}
        proxy = get_proxy()
        print(proxy)
        try:
            time.sleep(random.randint(5, 15))
            r = requests.get(url, headers=header, proxies=proxy, timeout=20)
            return r.text
            break
        except:
            continue

def get_data(html):
    soup = bs(html,'lxml')
    jobs = soup.find('div', id='jobs_list').find_all('div', id=re.compile(pattern))
    for job in jobs:
        specs = job.find_all('a', class_='skill')
        for spec in specs[1:]:
            skill = spec.text
            if skill in ('Старший (Senior)', 'Младший (Junior)', 'Средний (Middle)', 'Стажёр (Intern)'):
                continue
            skills_list.append(skill)


def get_xlsx(skills_list):
    wb = Workbook()
    ws = wb.active
    most_common = Counter(skills_list).most_common(6)
    for i in range(6):
        ws[f'A{i+1}'] = str(most_common[i][0])
        ws[f'B{i+1}'] = str(most_common[i][1])
    wb.save(f'moikrug_skills_{myskill}.xlsx')

def main():
    url = f'https://moikrug.ru/vacancies?skills%5B%5D={myskill_id}&type=all'
    get_data(get_html(url))
    for i in range(2, 6):
        try:
            url = f'https://moikrug.ru/vacancies?page={i}&skills%5B%5D=446&type=all'
            get_data(get_html(url))
        except:
            break
    get_xlsx(skills_list)


if __name__ == '__main__':
    main()